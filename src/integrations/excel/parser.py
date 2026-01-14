"""
Deterministic XLSX parser.

No guessing, no inference. Extracts only explicitly present data.
"""

import io
from typing import Any, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class ExcelParserError(Exception):
    """Excel parsing error."""
    pass


class ExcelParser:
    """Deterministic Excel parser."""

    def __init__(self):
        """Initialize parser."""
        pass

    def parse_xlsx(self, content: bytes) -> dict[str, list[list[Any]]]:
        """
        Parse XLSX file to raw data structure.

        Args:
            content: XLSX file content as bytes.

        Returns:
            Dictionary mapping sheet name to list of rows (each row is list of cells).

        Raises:
            ExcelParserError: If parsing fails.
        """
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        except Exception as e:
            raise ExcelParserError(f"Failed to load XLSX: {e}") from e

        sheets_data = {}

        for sheet_name in workbook.sheetnames:
            sheet: Worksheet = workbook[sheet_name]
            rows_data = []

            for row in sheet.iter_rows(values_only=True):
                rows_data.append(list(row))

            sheets_data[sheet_name] = rows_data

        return sheets_data

    def extract_text_from_xlsx(self, content: bytes) -> str:
        """
        Extract all text from XLSX as plain text.

        Args:
            content: XLSX file content as bytes.

        Returns:
            Plain text representation of all sheets.

        Raises:
            ExcelParserError: If parsing fails.
        """
        sheets_data = self.parse_xlsx(content)

        text_parts = []

        for sheet_name, rows in sheets_data.items():
            text_parts.append(f"[Sheet: {sheet_name}]")

            for row in rows:
                # Convert row to string, skip empty rows
                row_str = "\t".join(str(cell) if cell is not None else "" for cell in row)
                if row_str.strip():
                    text_parts.append(row_str)

            text_parts.append("")  # Empty line between sheets

        return "\n".join(text_parts)

    def find_header_row(self, rows: list[list[Any]], header_keywords: list[str]) -> Optional[int]:
        """
        Find header row by looking for specific keywords.

        Args:
            rows: List of rows.
            header_keywords: Keywords to search for in header (case-insensitive).

        Returns:
            Row index (0-based) or None if not found.
        """
        header_keywords_lower = [kw.lower() for kw in header_keywords]

        for i, row in enumerate(rows):
            # Convert row to lowercase strings
            row_values = [str(cell).lower() if cell is not None else "" for cell in row]

            # Check if all keywords are present in this row
            if all(any(kw in val for val in row_values) for kw in header_keywords_lower):
                return i

        return None

    def extract_column_data(
        self,
        rows: list[list[Any]],
        header_row_idx: int,
        column_name: str,
    ) -> list[Any]:
        """
        Extract data from a specific column.

        Args:
            rows: List of rows.
            header_row_idx: Index of header row.
            column_name: Name of column to extract (case-insensitive).

        Returns:
            List of values from that column (excluding header).
        """
        if header_row_idx >= len(rows):
            return []

        header_row = rows[header_row_idx]
        column_name_lower = column_name.lower()

        # Find column index
        column_idx = None
        for i, cell in enumerate(header_row):
            if cell is not None and column_name_lower in str(cell).lower():
                column_idx = i
                break

        if column_idx is None:
            return []

        # Extract data from that column
        data = []
        for row in rows[header_row_idx + 1 :]:
            if column_idx < len(row):
                data.append(row[column_idx])

        return data
