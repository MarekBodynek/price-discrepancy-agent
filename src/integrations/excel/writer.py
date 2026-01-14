"""
Excel report writer.

Writes CaseRow data to Excel file with mandatory column order.
"""

from datetime import date
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from src.core.models import CaseRow


class ExcelWriter:
    """Writes Excel reports."""

    # Column headers in mandatory order (per README)
    HEADERS = [
        "Unit (Store)",
        "EAN Code",
        "Document Creation Date",
        "Delivery Date",
        "Order Creation Date",
        "Supplier Price",
        "Internal (Own) Price",
        "Supplier Name",
        "Supplier Invoice Number",
        "Email Sender Address",
        "Email Link / Stable Reference",
        "Comments",
    ]

    def __init__(self):
        """Initialize writer."""
        pass

    def write_report(
        self,
        cases: list[CaseRow],
        output_path: str,
        date_from: date,
        date_to: date,
    ) -> None:
        """
        Write cases to Excel file.

        Args:
            cases: List of case rows.
            output_path: Output file path.
            date_from: Start date of run.
            date_to: End date of run.
        """
        # Create workbook
        wb = openpyxl.Workbook()
        ws: Worksheet = wb.active
        ws.title = "Price Discrepancies"

        # Write header row
        for col_idx, header in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)

        # Write data rows
        for row_idx, case in enumerate(cases, start=2):
            self._write_case_row(ws, row_idx, case)

        # Auto-adjust column widths
        for col_idx in range(1, len(self.HEADERS) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

        # Save workbook
        wb.save(output_path)

    def _write_case_row(self, ws: Worksheet, row_idx: int, case: CaseRow) -> None:
        """
        Write a single case row to worksheet.

        Args:
            ws: Worksheet.
            row_idx: Row index (1-based).
            case: Case row data.
        """
        # Column order must match HEADERS
        values = [
            case.unit_store,  # 1. Unit (Store)
            case.ean_code,  # 2. EAN Code
            self._format_date(case.document_creation_date),  # 3. Document Creation Date
            self._format_date(case.delivery_date),  # 4. Delivery Date
            self._format_date(case.order_creation_date),  # 5. Order Creation Date
            case.supplier_price,  # 6. Supplier Price
            case.internal_price,  # 7. Internal (Own) Price
            case.supplier_name or "",  # 8. Supplier Name
            case.supplier_invoice_number or "",  # 9. Supplier Invoice Number
            case.email_sender_address,  # 10. Email Sender Address
            case.email_link or "",  # 11. Email Link / Stable Reference
            case.comments,  # 12. Comments
        ]

        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    @staticmethod
    def _format_date(date_value: Optional[date]) -> str:
        """Format date for Excel."""
        if date_value is None:
            return ""
        return date_value.strftime("%Y-%m-%d")

    @staticmethod
    def generate_filename(date_from: date, date_to: date) -> str:
        """
        Generate filename for Excel report.

        Args:
            date_from: Start date.
            date_to: End date.

        Returns:
            Filename string.
        """
        if date_from == date_to:
            # Single date
            return f"Price_Discrepancies_{date_from.strftime('%Y-%m-%d')}.xlsx"
        else:
            # Date range
            return (
                f"Price_Discrepancies_{date_from.strftime('%Y-%m-%d')}_to_"
                f"{date_to.strftime('%Y-%m-%d')}.xlsx"
            )
