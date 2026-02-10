"""
Structured Excel extraction — preserves row-level EAN↔price↔store associations.

Parses Excel files row by row using openpyxl, mapping headers to internal fields.
Falls back to plain-text extraction if no recognizable headers are found.
"""

import io
import re
from typing import Any, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from src.core.models import DataSource, ExtractedData
from src.utils.text import (
    extract_dates,
    extract_suppliers,
    extract_invoice_numbers,
    find_date_by_keyword,
    is_valid_ean,
)


# Header mappings: internal field name → list of possible column header strings
# Supports Slovenian and English variants
HEADER_MAPPINGS: dict[str, list[str]] = {
    "ean": [
        "ean", "ean code", "ean koda", "ean-koda", "ean šifra",
        "črtna koda", "barcode", "bar code", "barkoda",
    ],
    "supplier_price": [
        "dobaviteljeva cena", "cena dobavitelja", "nabavna cena",
        "supplier price", "purchase price", "cost price", "cena",
        "vpc", "nc", "nab. cena", "nab cena",
    ],
    "internal_price": [
        "naša cena", "prodajna cena", "mpc", "maloprodajna cena",
        "internal price", "our price", "selling price", "retail price",
        "pc", "prod. cena", "prod cena",
    ],
    "store": [
        "enota", "trgovina", "poslovalnica", "unit", "store",
        "lokacija", "location", "poslovni prostor",
    ],
    "article_code": [
        "šifra", "šifra artikla", "artikel šifra", "article code",
        "item code", "product code", "art. šifra", "art šifra",
        "šifra izdelka", "code",
    ],
    "article_name": [
        "artikel", "naziv artikla", "ime artikla", "article name",
        "item name", "product name", "naziv", "opis", "description",
        "art. naziv", "art naziv",
    ],
    "quantity": [
        "količina", "kol", "qty", "quantity", "kol.",
    ],
    "unit_of_measure": [
        "enota mere", "em", "uom", "unit of measure", "me",
    ],
}


def _normalize_header(cell_value: Any) -> str:
    """Normalize a header cell value for matching."""
    if cell_value is None:
        return ""
    text = str(cell_value).strip().lower()
    # Remove extra whitespace
    text = re.sub(r'\s+', ' ', text)
    return text


def _find_header_row_and_mapping(
    rows: list[list[Any]],
) -> Optional[tuple[int, dict[str, int]]]:
    """
    Find the header row and map column indices to internal field names.

    Returns:
        Tuple of (header_row_index, {field_name: column_index}) or None.
    """
    # Only scan first 20 rows for headers
    scan_limit = min(20, len(rows))

    for row_idx in range(scan_limit):
        row = rows[row_idx]
        normalized = [_normalize_header(cell) for cell in row]

        # Try to map columns
        column_map: dict[str, int] = {}

        for field_name, possible_headers in HEADER_MAPPINGS.items():
            for col_idx, cell_text in enumerate(normalized):
                if not cell_text:
                    continue
                for header_variant in possible_headers:
                    if header_variant in cell_text or cell_text in header_variant:
                        # Prefer exact or longer match, avoid duplicates
                        if field_name not in column_map:
                            column_map[field_name] = col_idx
                        break

        # We need at least EAN column to consider this a valid header row
        # Also require at least one price column for meaningful extraction
        has_ean = "ean" in column_map
        has_price = "supplier_price" in column_map or "internal_price" in column_map

        if has_ean and has_price:
            return row_idx, column_map

        # Fallback: even just an EAN column with 2+ mapped fields is useful
        if has_ean and len(column_map) >= 2:
            return row_idx, column_map

    return None


def _extract_ean_from_cell(cell_value: Any) -> Optional[str]:
    """Extract and validate an EAN from a cell value."""
    if cell_value is None:
        return None

    text = str(cell_value).strip()

    # Remove common prefixes
    text = re.sub(r'^(ean|code|koda)\s*:?\s*', '', text, flags=re.IGNORECASE)

    # Extract digits only
    digits = re.sub(r'[^\d]', '', text)

    if len(digits) in (8, 13) and is_valid_ean(digits):
        return digits

    return None


def _extract_price_from_cell(cell_value: Any) -> Optional[float]:
    """Extract a price from a cell value."""
    if cell_value is None:
        return None

    # If already a number
    if isinstance(cell_value, (int, float)):
        val = float(cell_value)
        return val if val > 0 else None

    text = str(cell_value).strip()

    # Remove currency symbols and whitespace
    text = re.sub(r'[€$£\s]', '', text)
    text = text.replace('EUR', '').replace('USD', '').strip()

    # Handle comma as decimal separator
    if ',' in text and '.' in text:
        # 1.234,56 format → remove dots, replace comma with dot
        text = text.replace('.', '').replace(',', '.')
    elif ',' in text:
        text = text.replace(',', '.')

    try:
        val = float(text)
        return val if val > 0 else None
    except (ValueError, TypeError):
        return None


def extract_structured_from_excel(
    content: bytes,
    filename: str = "unknown.xlsx",
) -> Optional[ExtractedData]:
    """
    Extract structured data from Excel content, preserving row-level associations.

    Args:
        content: XLSX file content as bytes.
        filename: Original filename for source_details.

    Returns:
        ExtractedData with properly associated EAN↔price↔store data, or None.
    """
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        return None

    # Try each sheet
    for sheet_name in workbook.sheetnames:
        sheet: Worksheet = workbook[sheet_name]
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]

        if not rows:
            continue

        result = _find_header_row_and_mapping(rows)
        if result is None:
            continue

        header_row_idx, column_map = result

        # Parse data rows
        extracted = ExtractedData(
            source=DataSource.ATTACHMENT,
            source_details=f"Attachment: {filename} (structured)",
        )

        # Also try to extract metadata from rows above the header
        meta_text = ""
        for row in rows[:header_row_idx]:
            row_text = " ".join(str(c) for c in row if c is not None)
            meta_text += row_text + "\n"

        if meta_text:
            # Extract dates from metadata
            extracted.delivery_date = find_date_by_keyword(meta_text, "delivery|dostava|dobava")
            extracted.order_creation_date = find_date_by_keyword(meta_text, "order|naročilo|naročilnica")
            extracted.document_creation_date = find_date_by_keyword(meta_text, "document|created|date|datum|dokument")

            # Extract supplier info
            suppliers = extract_suppliers(meta_text)
            extracted.supplier_name = suppliers[0] if suppliers else None

            invoices = extract_invoice_numbers(meta_text)
            extracted.supplier_invoice_number = invoices[0] if invoices else None

        # Process data rows (after header)
        ean_col = column_map.get("ean")
        supplier_price_col = column_map.get("supplier_price")
        internal_price_col = column_map.get("internal_price")
        store_col = column_map.get("store")

        for row in rows[header_row_idx + 1:]:
            # Skip empty rows
            if all(c is None or str(c).strip() == "" for c in row):
                continue

            # Extract EAN
            ean = None
            if ean_col is not None and ean_col < len(row):
                ean = _extract_ean_from_cell(row[ean_col])

            if not ean:
                continue  # Skip rows without a valid EAN

            extracted.eans.append(ean)

            # Extract supplier price
            if supplier_price_col is not None and supplier_price_col < len(row):
                price = _extract_price_from_cell(row[supplier_price_col])
                if price is not None:
                    extracted.supplier_prices[ean] = price

            # Extract internal price
            if internal_price_col is not None and internal_price_col < len(row):
                price = _extract_price_from_cell(row[internal_price_col])
                if price is not None:
                    extracted.internal_prices[ean] = price

            # Extract store
            if store_col is not None and store_col < len(row):
                store_val = row[store_col]
                if store_val is not None and str(store_val).strip():
                    extracted.stores[ean] = str(store_val).strip()

        # If we found at least one EAN, return the structured extraction
        if extracted.eans:
            return extracted

    # No structured data found in any sheet
    return None
